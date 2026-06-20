import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "data", "exports")

SHOPEE_RED = "EE4D2D"
HEADER_TEXT = "FFFFFF"
ALT_ROW = "FFF5F3"


def _safe_filename(text: str) -> str:
    return re.sub(r"[^\w\-]", "_", text)


def export_to_excel(df: pd.DataFrame, keyword: str) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_kw = _safe_filename(keyword)[:30]
    filename = f"shopee_{safe_kw}_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Scraping"

    columns = {
        "item_id": "ID Produk",
        "name": "Nama Produk",
        "price_min": "Harga Min (IDR)",
        "price_max": "Harga Max (IDR)",
        "rating_star": "Rating",
        "sold": "Terjual",
        "liked_count": "Disukai",
        "shop_location": "Lokasi",
        "product_url": "URL Produk",
        "keyword": "Keyword",
        "scraped_at": "Waktu Scraping",
    }

    display_df = df[[c for c in columns if c in df.columns]].rename(columns=columns)

    header_fill = PatternFill("solid", fgColor=SHOPEE_RED)
    header_font = Font(bold=True, color=HEADER_TEXT, size=11)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Header row
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Data rows
    for row_idx, row in enumerate(display_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(r, col_idx).value or "")) for r in range(2, ws.max_row + 1)],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    return filepath
